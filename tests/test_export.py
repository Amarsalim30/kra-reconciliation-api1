import hashlib
import json
import io
import pytest
from datetime import date, datetime, timezone
from decimal import Decimal
from zipfile import ZipFile

from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database.base import Base
from app.database.database import get_db
from app.main import app
from app.schemas.invoice import Invoice, InvoiceSource
from app.domain.reconciliation_status import ReconciliationStatus
from app.domain.reconciliation_constants import (
    REMARK_MAP,
    STATUS_PRIORITY_VERSION,
    EXPORT_SCHEMA_VERSION,
)
from app.reporting.export_row import ReconciliationExportRow, to_export_rows
from app.reporting.utils import compute_sha256, build_status_counts
from app.reporting.errors import UnsupportedExportFormatError
from app.reporting.export_format import ExportFormat
from app.reporting.registry import ExportStrategyRegistry
from app.reporting.strategies.zip_strategy import ZipExporter
from app.reporting.export_filename_builder import ExportFilenameBuilder
from app.reporting.context import ExportContext
from app.services.summary_service import build_summary
from app.repositories.projections import ReconciliationProjection

# --- Test database setup ---
SQLALCHEMY_DATABASE_URL = "sqlite:///./test_export_db.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture(name="db_session", scope="function")
def fixture_db_session():
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture(name="client", scope="function")
def fixture_client(db_session):
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


@pytest.fixture(name="auth_headers", scope="function")
def fixture_auth_headers(client):
    register_payload = {
        "username": "export_tester",
        "password": "securepassword123",
        "email": "export_tester@example.com",
    }
    client.post("/api/v1/auth/register", json=register_payload)

    login_payload = {
        "username": "export_tester",
        "password": "securepassword123",
    }
    response = client.post("/api/v1/auth/login", json=login_payload)
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


# --- Unit tests for export_row ---

def _make_projection(**overrides) -> ReconciliationProjection:
    defaults = dict(
        cu_number="CU001",
        status=ReconciliationStatus.MATCH,
        amount_match=True,
        vat_match=True,
        date_match=True,
        sap_invoice_number="INV001",
        sap_partner_name="Test Partner",
        sap_pin="P001",
        sap_invoice_date=date(2026, 3, 1),
        sap_base_amount=Decimal("1000.00"),
        sap_vat_group="16",
        kra_invoice_number="INV001",
        kra_partner_name="Test Partner",
        kra_pin="P001",
        kra_invoice_date=date(2026, 3, 1),
        kra_base_amount=Decimal("1000.00"),
        kra_vat_group="16",
    )
    defaults.update(overrides)
    return ReconciliationProjection(**defaults)


def test_to_export_rows_populates_remark():
    projections = [_make_projection(status=ReconciliationStatus.AMOUNT_MISMATCH)]
    rows = to_export_rows(projections)
    assert len(rows) == 1
    assert rows[0].remark == REMARK_MAP[ReconciliationStatus.AMOUNT_MISMATCH]
    assert rows[0].status == ReconciliationStatus.AMOUNT_MISMATCH


def test_to_export_rows_preserves_all_fields():
    proj = _make_projection(
        cu_number="CU999",
        sap_invoice_number="SAP_INV",
        kra_invoice_number="KRA_INV",
        sap_pin="SAP_PIN",
        kra_pin="KRA_PIN",
    )
    rows = to_export_rows([proj])
    row = rows[0]
    assert row.cu_number == "CU999"
    assert row.sap_invoice_number == "SAP_INV"
    assert row.kra_invoice_number == "KRA_INV"
    assert row.sap_pin == "SAP_PIN"
    assert row.kra_pin == "KRA_PIN"


# --- Unit tests for summary_service ---

def test_build_summary_basic():
    rows = [
        to_export_rows([_make_projection(status=ReconciliationStatus.MATCH)])[0],
        to_export_rows([_make_projection(status=ReconciliationStatus.AMOUNT_MISMATCH)])[0],
    ]
    summary = build_summary(rows, total_sap=2, total_kra=2)
    assert summary.matches == 1
    assert summary.mismatches == 1
    assert summary.total_sap == 2
    assert summary.total_kra == 2
    assert summary.missing_in_sap == 0
    assert summary.missing_in_kra == 0
    assert summary.duplicate_cu == 0


def test_build_summary_empty_rows():
    summary = build_summary([], total_sap=0, total_kra=0)
    assert summary.matches == 0
    assert summary.total_sap == 0
    assert summary.match_percentage == 100.0


# --- Unit tests for SHA-256 ---

def test_sha256_same_rows_same_hash():
    rows = to_export_rows([
        _make_projection(cu_number="CU001", status=ReconciliationStatus.MATCH),
        _make_projection(cu_number="CU002", status=ReconciliationStatus.AMOUNT_MISMATCH),
    ])
    h1 = compute_sha256(rows, EXPORT_SCHEMA_VERSION)
    h2 = compute_sha256(rows, EXPORT_SCHEMA_VERSION)
    assert h1 == h2
    assert len(h1) == 64  # SHA-256 hex digest


def test_sha256_different_order_same_hash():
    rows1 = to_export_rows([
        _make_projection(cu_number="CU001", status=ReconciliationStatus.MATCH),
        _make_projection(cu_number="CU002", status=ReconciliationStatus.AMOUNT_MISMATCH),
    ])
    rows2 = to_export_rows([
        _make_projection(cu_number="CU002", status=ReconciliationStatus.AMOUNT_MISMATCH),
        _make_projection(cu_number="CU001", status=ReconciliationStatus.MATCH),
    ])
    h1 = compute_sha256(rows1, EXPORT_SCHEMA_VERSION)
    h2 = compute_sha256(rows2, EXPORT_SCHEMA_VERSION)
    assert h1 == h2


def test_sha256_includes_schema_version():
    rows = to_export_rows([
        _make_projection(cu_number="CU001", status=ReconciliationStatus.MATCH),
    ])
    h1 = compute_sha256(rows, "1.0")
    h2 = compute_sha256(rows, "2.0")
    assert h1 != h2


def test_sha256_changes_on_amount_change():
    rows1 = to_export_rows([_make_projection(
        cu_number="CU001",
        sap_base_amount=Decimal("1000.00"),
    )])
    rows2 = to_export_rows([_make_projection(
        cu_number="CU001",
        sap_base_amount=Decimal("2000.00"),
    )])
    h1 = compute_sha256(rows1, EXPORT_SCHEMA_VERSION)
    h2 = compute_sha256(rows2, EXPORT_SCHEMA_VERSION)
    assert h1 != h2


# --- Unit tests for status counts ---

def testbuild_status_counts():
    rows = to_export_rows([
        _make_projection(status=ReconciliationStatus.MATCH),
        _make_projection(status=ReconciliationStatus.MATCH),
        _make_projection(status=ReconciliationStatus.AMOUNT_MISMATCH),
    ])
    counts = build_status_counts(rows)
    assert counts["Match"] == 2
    assert counts["Amount Mismatch"] == 1


# --- Unit tests for remark derivation ---

def test_remark_derived_from_current_map():
    for status in ReconciliationStatus:
        row = to_export_rows([_make_projection(status=status)])[0]
        assert row.remark == REMARK_MAP[status]


# --- Unit tests for filename builder ---

def test_filename_builder():
    from app.models.reconciliation_session import ReconciliationSession
    from app.schemas.invoice import ReconciliationType

    builder = ExportFilenameBuilder()
    # Create a minimal mock session
    class MockSession:
        session_type = ReconciliationType.SALES
        from_date = date(2026, 3, 1)
        to_date = date(2026, 3, 31)

    ts = datetime(2026, 7, 10, 17, 47, 0, 382000, tzinfo=timezone.utc)
    filename = builder.build(MockSession(), ts, ExportFormat.ZIP)
    assert filename == "Sales_Reconciliation_2026-03-01_to_2026-03-31_20260710T174700.382Z.zip"
    assert "20260710T" in filename
    assert filename.endswith(".zip")


# --- Unit tests for registry ---

def test_registry_get_returns_strategy():
    registry = ExportStrategyRegistry()
    registry.register(ExportFormat.ZIP, ZipExporter())
    strategy = registry.get(ExportFormat.ZIP)
    assert isinstance(strategy, ZipExporter)


def test_registry_raises_for_unknown_format():
    registry = ExportStrategyRegistry()
    with pytest.raises(UnsupportedExportFormatError):
        registry.get(ExportFormat.ZIP)  # not registered


# --- Unit tests for ExportContext ---

def test_export_context_rejects_naive_datetime():
    from pydantic import ValidationError
    with pytest.raises(ValueError, match="timezone-aware"):
        ExportContext(
            generated_by="test",
            app_version="1.0.0",
            generated_at=datetime(2026, 1, 1),  # naive
        )


def test_export_context_accepts_aware_datetime():
    ctx = ExportContext(
        generated_by="test",
        app_version="1.0.0",
        generated_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
    )
    assert ctx.export_version == EXPORT_SCHEMA_VERSION


# --- Integration tests ---

def _setup_compared_session(client, auth_headers) -> str:
    """Helper: create a session with compared results, return session_id."""
    load_res = client.get("/api/v1/sales?from=2026-03-01&to=2026-03-30", headers=auth_headers)
    assert load_res.status_code == 200
    session_id = load_res.json()["session_id"]

    csv_content = (
        b"Pin Number,Customer Name,Invoice Number,Invoice Date,CU Number,VAT Group,Base Amount\n"
        b"P051393568M,Autoports Freight Terminals Limited,IN1080,02/03/2026,|0190439340000000455,16,1118894.84\n"
        b"P051137818X,GRAIN INDUSTRIES LIMITED,IN1081,11/03/2026,|0190439340000000456,16,3977701.88\n"
    )
    files = [("files", ("SEC_B.csv", csv_content, "text/csv"))]
    upload_res = client.post(f"/api/v1/sales/upload?session_id={session_id}", headers=auth_headers, files=files)
    assert upload_res.status_code == 200

    compare_res = client.post(
        "/api/v1/reconciliation/compare",
        headers=auth_headers,
        json={"session_id": session_id},
    )
    assert compare_res.status_code == 200
    return session_id


def test_export_returns_zip(client, auth_headers):
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/zip"
    assert len(res.content) > 0


def test_export_zip_structure(client, auth_headers):
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)
    assert res.status_code == 200

    with ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
        assert "_metadata.json" in names
        assert "01 Summary.xlsx" in names
        assert "02 Exceptions.xlsx" in names
        assert "03 Matches.xlsx" in names


def test_export_metadata_json_fields(client, auth_headers):
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)

    with ZipFile(io.BytesIO(res.content)) as zf:
        export_json = json.loads(zf.read("_metadata.json"))
        assert export_json["schema_version"] == "2.0"
        assert "status_priority_version" in export_json
        assert "record_counts" in export_json
        assert "checksum" in export_json
        assert "generated_at" in export_json
        assert "application_version" in export_json
        assert export_json["session_id"] == session_id


def test_export_requires_compared_session(client, auth_headers):
    # Load SAP but don't compare
    load_res = client.get("/api/v1/sales?from=2026-03-01&to=2026-03-30", headers=auth_headers)
    session_id = load_res.json()["session_id"]

    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)
    assert res.status_code == 400
    assert "compared" in res.json()["detail"].lower()


def test_export_requires_auth(client):
    res = client.get("/api/v1/reconciliation/dummy/export?format=zip")
    assert res.status_code == 401


def test_export_unknown_format_returns_400(client, auth_headers):
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=csv", headers=auth_headers)
    assert res.status_code == 422  # FastAPI validation error for invalid enum


def test_export_filename_is_timestamped(client, auth_headers):
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)
    disposition = res.headers.get("content-disposition", "")
    assert "Sales_Reconciliation" in disposition
    assert ".zip" in disposition
    # Millisecond timestamp pattern: T followed by digits then Z
    assert "T" in disposition
    assert "Z.zip" in disposition


def test_export_empty_session(client, auth_headers, db_session):
    """Export an empty result set — should return HTTP 200 with README.txt."""
    from app.models.reconciliation_session import ReconciliationSession as ReconSession, SessionReconciliationResult
    from app.models.user import User

    user = db_session.query(User).filter(User.username == "export_tester").first()

    session = ReconSession(
        user_id=user.id,
        from_date=date(2026, 3, 1),
        to_date=date(2026, 3, 31),
        is_compared=True,
        comparison_results={"summary": {
            "total_sap": 0, "total_kra": 0,
            "matches": 0, "missing_in_sap": 0, "missing_in_kra": 0,
            "mismatches": 0, "duplicate_cu": 0,
            "match_percentage": 100.0, "completion_percentage": 100.0,
            "mismatch_stats": {"amount": 0, "vat": 0, "date": 0},
        }},
    )
    db_session.add(session)
    db_session.commit()

    res = client.get(f"/api/v1/reconciliation/{session.id}/export?format=zip", headers=auth_headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/zip"

    with ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
        assert "_metadata.json" in names
        assert "01 Summary.xlsx" in names
        assert "02 Exceptions.xlsx" not in names
        assert "03 Matches.xlsx" not in names


def test_export_large_dataset_correctness(client, auth_headers, db_session):
    """Insert 100 rows directly into DB and verify export preserves count."""
    from app.models.reconciliation_session import ReconciliationSession, SessionReconciliationResult

    # Create a session
    session = ReconciliationSession(
        user_id=1,
        from_date=date(2026, 3, 1),
        to_date=date(2026, 3, 31),
        is_compared=True,
        comparison_results={"summary": {
            "total_sap": 100, "total_kra": 100,
            "matches": 50, "missing_in_sap": 10, "missing_in_kra": 10,
            "mismatches": 30, "duplicate_cu": 0,
            "match_percentage": 50.0, "completion_percentage": 50.0,
            "mismatch_stats": {"amount": 10, "vat": 10, "date": 10},
        }},
    )
    db_session.add(session)
    db_session.flush()

    # Insert 100 results
    for i in range(100):
        status = ReconciliationStatus.MATCH if i % 2 == 0 else ReconciliationStatus.AMOUNT_MISMATCH
        db_session.add(SessionReconciliationResult(
            session_id=session.id,
            row_number=i + 1,
            cu_number=f"CU{i:04d}",
            status=status,
            amount_match=(status == ReconciliationStatus.MATCH),
            vat_match=True,
            date_match=True,
            sap_invoice_number=f"INV{i:04d}",
            sap_partner_name=f"Partner {i}",
            sap_pin=f"P{i:04d}",
            sap_invoice_date=date(2026, 3, 1),
            sap_base_amount=Decimal("1000.00"),
            sap_vat_group="16",
            kra_invoice_number=f"INV{i:04d}",
            kra_partner_name=f"Partner {i}",
            kra_pin=f"P{i:04d}",
            kra_invoice_date=date(2026, 3, 1),
            kra_base_amount=Decimal("1000.00"),
            kra_vat_group="16",
        ))
    db_session.commit()

    # Register a user for auth
    client.post("/api/v1/auth/register", json={
        "username": "large_export_user",
        "password": "securepassword123",
        "email": "large@example.com",
    })
    login_res = client.post("/api/v1/auth/login", json={
        "username": "large_export_user",
        "password": "securepassword123",
    })
    token = login_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Update session user_id to match
    session.user_id = login_res.json().get("user_id", 1)
    db_session.commit()

    # Actually we need to set the user_id correctly. Let's query it.
    from app.models.user import User
    user = db_session.query(User).filter(User.username == "large_export_user").first()
    session.user_id = user.id
    db_session.commit()

    res = client.get(f"/api/v1/reconciliation/{session.id}/export?format=zip", headers=headers)
    assert res.status_code == 200

    with ZipFile(io.BytesIO(res.content)) as zf:
        export_json = json.loads(zf.read("_metadata.json"))
        assert export_json["record_counts"]["matches"] == 50
        assert export_json["record_counts"]["exceptions"] == 50


def test_export_needs_review_excludes_matches(client, auth_headers):
    """The Needs Review sheet should not contain MATCH rows."""
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)

    with ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
        assert "02 Exceptions.xlsx" in names
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(zf.read("02 Exceptions.xlsx")))
        assert "Matches" not in wb.sheetnames


def test_export_matches_compact_no_sap_kra_duplicates(client, auth_headers):
    """The Matches sheet should not have duplicated SAP/KRA columns (compact layout)."""
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)

    with ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
        assert "03 Matches.xlsx" in names
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(zf.read("03 Matches.xlsx")))
        ws = wb["Matches"]
        headers = [cell.value for cell in ws[1]]
        # Compact: should NOT have KRA PIN/Partner columns
        assert "KRA PIN" not in headers
        assert "KRA Partner" not in headers


def test_export_missing_summary_handling(client, auth_headers, db_session):
    from app.models.reconciliation_session import ReconciliationSession
    from app.models.user import User

    user = db_session.query(User).filter(User.username == "export_tester").first()
    
    # Create session marked as compared, but comparison_results is empty/missing summary
    session = ReconciliationSession(
        user_id=user.id,
        from_date=date(2026, 3, 1),
        to_date=date(2026, 3, 31),
        is_compared=True,
        comparison_results={},  # missing summary!
    )
    db_session.add(session)
    db_session.commit()

    res = client.get(f"/api/v1/reconciliation/{session.id}/export?format=zip", headers=auth_headers)
    assert res.status_code == 500
    assert "Unable to generate export due to an internal reconciliation state error." in res.json()["detail"]


def test_worksheet_order_preservation(client, auth_headers):
    # Setup session with some exceptions
    session_id = _setup_compared_session(client, auth_headers)
    res = client.get(f"/api/v1/reconciliation/{session_id}/export?format=zip", headers=auth_headers)
    assert res.status_code == 200

    with ZipFile(io.BytesIO(res.content)) as zf:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(zf.read("02 Exceptions.xlsx")))
        
        # Defined order
        defined_order = [
            "Missing in SAP",
            "Missing in KRA",
            "Amount Mismatch",
            "VAT Mismatch",
            "Duplicate CU",
            "Multiple Issues"
        ]
        
        # Verify the actual sheets in wb.sheetnames are in correct relative order
        actual_sheets = wb.sheetnames
        indexes = [defined_order.index(name) for name in actual_sheets if name in defined_order]
        assert indexes == sorted(indexes), f"Sheet order is not preserved: {actual_sheets}"
