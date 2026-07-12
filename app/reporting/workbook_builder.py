from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.reporting.context import ExportContext
from app.reporting.export_row import ReconciliationExportRow
from app.reporting.styles import (
    AMOUNT_FORMAT, HEADER_FILL, HEADER_FONT, SUBTITLE_FONT,
    TITLE_FONT, VALUE_FONT, WARNING_FILL,
)
from app.reporting.workbook_artifact import WorkbookArtifact
from app.schemas.reconciliation import ReconciliationSummary

from app.models.reconciliation_session import ReconciliationSession
from datetime import timezone
from app.reporting.sheet_definitions import WORKBOOK_DEFINITIONS, WorkbookDefinition, SheetDefinition, SheetColumn
from app.domain.reconciliation_status import ReconciliationStatus


def build_all(
    rows: list[ReconciliationExportRow],
    summary: ReconciliationSummary,
    context: ExportContext,
    session: ReconciliationSession,
) -> list[WorkbookArtifact]:
    """Build all workbook artifacts: summary + exception/matches workbooks."""
    artifacts: list[WorkbookArtifact] = []

    # 1. Summary workbook
    artifacts.append(_build_summary_workbook(rows, summary, context, session))

    # 2. Detail workbooks
    for workbook_def in WORKBOOK_DEFINITIONS:
        # Check if there is at least one sheet in this workbook that will have data
        has_data = False
        for sheet_def in workbook_def.sheets:
            sheet_rows = [r for r in rows if sheet_def.statuses and r.status in sheet_def.statuses]
            if len(sheet_rows) > 0:
                has_data = True
                break
        
        if has_data:
            artifacts.append(_build_detail_workbook(rows, workbook_def, context))

    return artifacts


def _build_summary_workbook(
    rows: list[ReconciliationExportRow],
    summary: ReconciliationSummary,
    context: ExportContext,
    session: ReconciliationSession,
) -> WorkbookArtifact:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Reconciliation Summary"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    # Generated At
    ws.cell(row=3, column=1, value="Generated At:").font = SUBTITLE_FONT
    ws.cell(row=4, column=1, value=context.generated_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")).font = VALUE_FONT

    # Date Range
    ws.cell(row=6, column=1, value="Date Range:").font = SUBTITLE_FONT
    ws.cell(row=7, column=1, value=f"{session.from_date} to {session.to_date}").font = VALUE_FONT

    # Session Type
    ws.cell(row=9, column=1, value="Session Type:").font = SUBTITLE_FONT
    ws.cell(row=10, column=1, value=session.session_type.value.title()).font = VALUE_FONT

    # Record Counts section
    ws.cell(row=13, column=1, value="Record Counts").font = TITLE_FONT
    ws.merge_cells("A13:D13")

    needs_review = summary.total_reconciled_rows - summary.matches
    match_rate = summary.match_percentage if summary.total_reconciled_rows > 0 else 0.0

    counts = [
        ("SAP Invoices", summary.total_sap),
        ("KRA Invoices", summary.total_kra),
        ("Matched", summary.matches),
        ("Needs Review", needs_review),
        ("Total Reconciled Rows", summary.total_reconciled_rows),
        ("Match Rate", f"{match_rate:.1f}%"),
    ]

    row = 15
    for label, value in counts:
        ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
        ws.cell(row=row, column=2, value=value).font = VALUE_FONT
        row += 1

    # Needs Review Breakdown (only if there are exceptions)
    if needs_review > 0:
        row += 1
        ws.cell(row=row, column=1, value="Needs Review Breakdown").font = TITLE_FONT
        ws.cell(row=row+1, column=1, value="(Only categories requiring action)").font = SUBTITLE_FONT
        row += 2

        # Count actual statuses directly from rows to ensure consistency
        breakdown_counts = [
            ("Missing in SAP", sum(1 for r in rows if r.status == ReconciliationStatus.MISSING_IN_SAP)),
            ("Missing in KRA", sum(1 for r in rows if r.status == ReconciliationStatus.MISSING_IN_KRA)),
            ("Amount Mismatch", sum(1 for r in rows if r.status == ReconciliationStatus.AMOUNT_MISMATCH)),
            ("VAT Mismatch", sum(1 for r in rows if r.status == ReconciliationStatus.VAT_MISMATCH)),
            ("Duplicate CU", sum(1 for r in rows if r.status == ReconciliationStatus.DUPLICATE_SOURCE_KEY)),
            ("Multiple Issues", sum(1 for r in rows if r.status == ReconciliationStatus.MULTIPLE_MISMATCHES)),
        ]

        for label, val in breakdown_counts:
            if val > 0:
                ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
                ws.cell(row=row, column=2, value=val).font = VALUE_FONT
                row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return WorkbookArtifact(
        zip_path="01 Summary.xlsx",
        filename="01 Summary.xlsx",
        content=buf.read(),
    )


def _build_detail_workbook(
    all_rows: list[ReconciliationExportRow],
    workbook_def: WorkbookDefinition,
    context: ExportContext,
) -> WorkbookArtifact:
    wb = Workbook()
    
    # Remove default sheet to avoid empty "Sheet" at the end
    default_sheet = wb.active

    for i, sheet_def in enumerate(workbook_def.sheets):
        # Filter rows for this sheet
        rows = [r for r in all_rows if sheet_def.statuses and r.status in sheet_def.statuses]

        if not rows:
            continue

        ws = wb.create_sheet(title=sheet_def.title)
        
        # Header row
        headers = [col.header for col in sheet_def.columns]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 32

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_def in enumerate(sheet_def.columns, 1):
                value = getattr(row_data, col_def.attr)
                cell = ws.cell(row=row_idx, column=col_idx)

                if value is None:
                    cell.value = ""
                elif col_def.is_amount:
                    cell.value = float(value)
                    cell.number_format = AMOUNT_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                elif col_def.is_date:
                    cell.value = value.isoformat() if hasattr(value, "isoformat") else str(value)
                    cell.alignment = Alignment(horizontal="center")
                elif col_def.is_bool:
                    cell.value = "Yes" if value else "No"
                    cell.alignment = Alignment(horizontal="center")
                    if not value:
                        cell.fill = WARNING_FILL
                elif col_def.is_cu:
                    cell.value = str(value)
                    cell.number_format = "@"  # text format
                else:
                    cell.value = str(value)

                cell.font = VALUE_FONT

        # Apply column widths
        for col_idx, col_def in enumerate(sheet_def.columns, 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = col_def.width

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-filter and Excel Table (only if there are rows)
        if rows:
            last_col = get_column_letter(len(headers))
            table_ref = f"A1:{last_col}{len(rows) + 1}"
            from openpyxl.worksheet.table import Table, TableStyleInfo
            # Sanitize table name for Excel
            safe_name = f"tbl_{sheet_def.title.replace(' ', '_').lower()}"
            table = Table(displayName=safe_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

    # Remove the initially created empty sheet after all other sheets are added
    if default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return WorkbookArtifact(
        zip_path=workbook_def.filename,
        filename=workbook_def.filename,
        content=buf.read(),
    )
