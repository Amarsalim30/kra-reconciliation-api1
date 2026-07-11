from openpyxl.styles import Alignment, Font, PatternFill, numbers

# ---------------------------------------------------------------------------
# Fills
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
MATCH_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="4B5563")
VALUE_FONT = Font(name="Calibri", size=11)

# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------
AMOUNT_FORMAT = '#,##0.00'

# ---------------------------------------------------------------------------
# Column widths (approximate character widths for openpyxl)
# ---------------------------------------------------------------------------
CU_COL_WIDTH = 22
TEXT_COL_WIDTH = 30
AMOUNT_COL_WIDTH = 18
DATE_COL_WIDTH = 14
BOOL_COL_WIDTH = 12
